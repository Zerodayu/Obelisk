import asyncio
from pathlib import Path
from typing import Any, List, Dict
from openpyxl import load_workbook
from openpyxl.utils.cell import column_index_from_string
from openpyxl.worksheet.worksheet import Worksheet

from app.core.exceptions import InvalidTemplate, InvalidWorkbook, MissingWorksheet
from app.core.logging import logger
from app.etl.abstracts import Extractor
from app.schemas.class_record import ClassRecordHeader, RawScoreRecord

from .. import etl_const


class ExcelExtractor(Extractor):
    """Extracts data from a standard JMCFI class-record Excel workbook."""

    async def extract(self, source: Any) -> tuple[ClassRecordHeader, list[RawScoreRecord], list[dict[str, Any]]]:
        """
        Main entrypoint to extract all data from a given workbook source.
        """
        file_path = self._resolve_file_path(source)
        return await asyncio.to_thread(self._extract_sync, file_path)

    def _extract_sync(self, file_path: str) -> tuple[ClassRecordHeader, list[RawScoreRecord], list[dict[str, Any]]]:
        """Synchronous wrapper for all extraction operations."""
        try:
            workbook = load_workbook(file_path, data_only=True)
        except Exception as exc:
            raise InvalidWorkbook(file_path=file_path, underlying_error=str(exc))

        db_sheet = self._require_sheet(workbook, etl_const.SheetNames.DATABASE)
        exam_sheet = self._require_sheet(workbook, etl_const.SheetNames.EXAM)
        cover_sheet = self._require_sheet(workbook, etl_const.SheetNames.COVERPAGE)
        output_sheet = workbook[etl_const.SheetNames.OUTPUT] if etl_const.SheetNames.OUTPUT in workbook.sheetnames else None

        self._validate_template(db_sheet, "B12")
        self._validate_template(exam_sheet, "B18")
        if output_sheet is not None:
            self._validate_template(output_sheet, "B18")

        header = self._build_header(db_sheet, cover_sheet)
        clo_plo_mapping = self._extract_clo_plo_mapping(cover_sheet)

        db_students = self._read_roster(db_sheet, start_row=etl_const.Roster.DATABASE_START_ROW)
        records: list[RawScoreRecord] = []

        records.extend(self._extract_database_block(sheet=db_sheet, students=db_students, grading_period=etl_const.GradingPeriod.PRELIM, columns=etl_const.DatabaseSheet.PRELIM_COLS))
        records.extend(self._extract_database_block(sheet=db_sheet, students=db_students, grading_period=etl_const.GradingPeriod.MIDTERM, columns=etl_const.DatabaseSheet.MIDTERM_COLS))
        records.extend(self._extract_database_block(sheet=db_sheet, students=db_students, grading_period=etl_const.GradingPeriod.FINAL, columns=etl_const.DatabaseSheet.FINAL_COLS))

        db_students_by_name = {self._normalize_name(student["student_name"]): student for student in db_students if student["student_name"]}
        records.extend(self._extract_exam_sheet(exam_sheet, db_students_by_name))

        if output_sheet is not None:
            records.extend(self._extract_output_sheet(output_sheet, db_students_by_name))

        return header, records, clo_plo_mapping

    def _extract_clo_plo_mapping(self, sheet: Worksheet) -> List[Dict[str, Any]]:
        """
        Extracts the CLO-PLO correlation mapping table from the COVERPAGE.
        """
        try:
            self._validate_template(sheet, etl_const.CloPlo.TABLE_HEADER_CELL, expected=etl_const.CloPlo.TABLE_HEADER_VALUE)
        except InvalidTemplate:
            logger.warning("clo_plo_mapping_not_found", sheet=sheet.title, reason=f"Header '{etl_const.CloPlo.TABLE_HEADER_VALUE}' not found at {etl_const.CloPlo.TABLE_HEADER_CELL}.")
            return []
        
        mapping = []
        plo_headers = {}
        for col_idx in range(2, sheet.max_column + 1):
            plo_code = self._as_optional_string(sheet.cell(row=etl_const.CloPlo.PLO_HEADER_ROW, column=col_idx).value)
            if not plo_code: break
            plo_headers[col_idx] = plo_code
        
        for row_idx in range(etl_const.CloPlo.FIRST_CLO_ROW, sheet.max_row + 1):
            clo_code = self._as_optional_string(sheet.cell(row=row_idx, column=1).value)
            if not clo_code or clo_code.strip().upper() == etl_const.CloPlo.END_OF_TABLE_SENTINEL: break
            for col_idx, plo_code in plo_headers.items():
                correlation = sheet.cell(row=row_idx, column=col_idx).value
                if correlation is not None and str(correlation).strip() != "":
                    mapping.append({"clo_code": clo_code, "plo_code": plo_code, "correlation_strength": self._as_int(correlation)})
        
        logger.info("clo_plo_mapping_extracted", count=len(mapping))
        return mapping

    @staticmethod
    def _resolve_file_path(source: Any) -> str:
        """Finds a valid file path from various possible source types."""
        if isinstance(source, dict):
            path = source.get("file_path") or source.get("path")
            if path: return str(path)
        if hasattr(source, "file_path"):
            path = getattr(source, "file_path")
            if path: return str(path)
        if isinstance(source, (str, Path)):
            return str(source)
        raise InvalidWorkbook(file_path=str(source), underlying_error="Invalid source type")

    def _require_sheet(self, workbook: Any, sheet_name: str) -> Worksheet:
        """Finds a sheet by name in the workbook, raising a detailed error if not found."""
        if sheet_name not in workbook.sheetnames:
            raise MissingWorksheet(expected_sheet_name=sheet_name, available_sheets=list(workbook.sheetnames))
        return workbook[sheet_name]

    def _validate_template(self, sheet: Worksheet, header_cell: str, expected: str = "STUDENT NAME") -> None:
        """Checks for a specific header value in a cell to validate the template version."""
        header_value = self._normalize_text(sheet[header_cell].value)
        if header_value != expected:
            raise InvalidTemplate(sheet_name=sheet.title, cell=header_cell, expected=expected, found=str(header_value))

    def _build_header(self, db_sheet: Worksheet, cover_sheet: Worksheet) -> ClassRecordHeader:
        semester_year = self._as_string(db_sheet[etl_const.HeaderData.SEMESTER_YEAR].value)
        course_type = self._as_string(db_sheet[etl_const.HeaderData.COURSE_TYPE].value)
        no_of_students = self._as_int(db_sheet[etl_const.HeaderData.NO_OF_STUDENTS].value)
        threshold = self._as_float(db_sheet[etl_const.HeaderData.THRESHOLD].value)
        
        weights: dict[str, float] = {}
        d5 = self._as_string(db_sheet[etl_const.HeaderData.WEIGHT_COMPONENT_1_NAME].value)
        if d5: weights[d5] = self._as_float(db_sheet[etl_const.HeaderData.WEIGHT_COMPONENT_1_VALUE].value)
        e5 = self._as_string(db_sheet[etl_const.HeaderData.WEIGHT_COMPONENT_2_NAME].value)
        if e5: weights[e5] = self._as_float(db_sheet[etl_const.HeaderData.WEIGHT_COMPONENT_2_VALUE].value)
        f5 = self._as_string(db_sheet[etl_const.HeaderData.WEIGHT_COMPONENT_3_NAME].value)
        if f5: weights[f5] = self._as_float(db_sheet[etl_const.HeaderData.WEIGHT_COMPONENT_3_VALUE].value)
        
        return ClassRecordHeader(
            course_code=self._coalesce_optional(db_sheet[etl_const.HeaderData.COURSE_CODE].value, self._find_cover_value(cover_sheet, "Course Code:")),
            course_title=self._coalesce_optional(db_sheet[etl_const.HeaderData.COURSE_TITLE].value, self._find_cover_value(cover_sheet, "Course Title:")),
            course_type=course_type,
            section=self._coalesce_optional(db_sheet[etl_const.HeaderData.SECTION].value, self._find_cover_value(cover_sheet, "Section:")),
            semester_year=semester_year,
            instructor_name=self._coalesce_optional(db_sheet[etl_const.HeaderData.INSTRUCTOR_NAME].value, self._find_cover_value(cover_sheet, "Instructor's Name")),
            no_of_students=no_of_students,
            threshold=threshold,
            grading_system=self._coalesce_optional(db_sheet[etl_const.HeaderData.GRADING_SYSTEM].value, self._find_cover_value(cover_sheet, "GRADING SYSTEM")),
            workbook_configured_weights_unused=weights if weights else None,
        )

    def _read_roster(self, sheet: Worksheet, start_row: int) -> list[dict[str, str | int | None]]:
        students: list[dict[str, str | int | None]] = []
        row = start_row
        while True:
            student_id = self._as_optional_string(sheet[f"A{row}"].value)
            student_name = self._as_optional_string(sheet[f"B{row}"].value)
            if not student_id and not student_name:
                break
            if student_name:
                students.append({"student_id": student_id, "student_name": student_name, "row": row})
            row += 1
        return students

    def _extract_database_block(self, sheet: Worksheet, students: list[dict[str, str | int | None]], grading_period: str, columns: list[str]) -> list[RawScoreRecord]:
        records: list[RawScoreRecord] = []
        for col in columns:
            max_score = self._as_optional_float(sheet[f"{col}{etl_const.DatabaseSheet.MAX_SCORE_ROW}"].value)
            clo_code = self._as_optional_string(sheet[f"{col}{etl_const.DatabaseSheet.CLO_CODE_ROW}"].value)
            if max_score is None or clo_code is None: continue
            
            assessment_category = self._as_string(sheet[f"{col}{etl_const.DatabaseSheet.ASSESSMENT_CATEGORY_ROW}"].value)
            assessment_no = self._as_int(sheet[f"{col}{etl_const.DatabaseSheet.ASSESSMENT_NO_ROW}"].value)
            activity_name = self._as_optional_string(sheet[f"{col}{etl_const.DatabaseSheet.ACTIVITY_NAME_ROW}"].value)
            col_idx = column_index_from_string(col)
            
            for student in students:
                raw_score = self._as_optional_float(sheet.cell(row=student["row"], column=col_idx).value)
                records.append(RawScoreRecord(student_id=student["student_id"], student_name=student["student_name"] or "", grading_period=grading_period, assessment_category=assessment_category, assessment_no=assessment_no, clo_code=clo_code, activity_name=activity_name, max_score=max_score, raw_score=raw_score))
        return records

    def _extract_exam_sheet(self, sheet: Worksheet, db_students_by_name: dict[str, dict[str, str | int | None]]) -> list[RawScoreRecord]:
        exam_students = self._read_roster(sheet, start_row=etl_const.Roster.EXAM_AND_OUTPUT_START_ROW)
        student_lookup = self._merge_roster_by_name(exam_students, db_students_by_name)
        records: list[RawScoreRecord] = []
        
        records.extend(self._extract_exam_columns(sheet=sheet, students=student_lookup, grading_period=etl_const.GradingPeriod.PRELIM, columns=etl_const.ExamSheet.PRELIM_COLS, activity_name="Prelim Exam"))
        records.extend(self._extract_exam_columns(sheet=sheet, students=student_lookup, grading_period=etl_const.GradingPeriod.MIDTERM, columns=etl_const.ExamSheet.MIDTERM_COLS, activity_name="Midterm Exam"))
        records.extend(self._extract_exam_columns(sheet=sheet, students=student_lookup, grading_period=etl_const.GradingPeriod.FINAL, columns=etl_const.ExamSheet.FINAL_COLS, activity_name="Final Exam"))
        
        return records

    def _extract_exam_columns(self, sheet: Worksheet, students: list[dict[str, str | int | None]], grading_period: str, columns: list[str], activity_name: str) -> list[RawScoreRecord]:
        records: list[RawScoreRecord] = []
        for idx, col in enumerate(columns, start=1):
            max_score = self._as_optional_float(sheet[f"{col}{etl_const.ExamSheet.MAX_SCORE_ROW}"].value)
            if max_score is None: continue
            
            clo_code = self._as_optional_string(sheet[f"{col}{etl_const.ExamSheet.CLO_CODE_ROW}"].value)
            if clo_code is None: continue
            
            col_idx = column_index_from_string(col)
            for student in students:
                raw_score = self._as_optional_float(sheet.cell(row=student["row"], column=col_idx).value)
                records.append(RawScoreRecord(student_id=student["student_id"], student_name=student["student_name"] or "", grading_period=grading_period, assessment_category="EXAM", assessment_no=idx, clo_code=clo_code, activity_name=activity_name, max_score=max_score, raw_score=raw_score))
        return records

    def _extract_output_sheet(self, sheet: Worksheet, db_students_by_name: dict[str, dict[str, str | int | None]]) -> list[RawScoreRecord]:
        output_students = self._read_roster(sheet, start_row=etl_const.Roster.EXAM_AND_OUTPUT_START_ROW)
        students = self._merge_roster_by_name(output_students, db_students_by_name)
        records: list[RawScoreRecord] = []
        
        period_columns = {
            etl_const.GradingPeriod.PRELIM: etl_const.OutputSheet.PRELIM_COLS,
            etl_const.GradingPeriod.MIDTERM: etl_const.OutputSheet.MIDTERM_COLS,
            etl_const.GradingPeriod.FINAL: etl_const.OutputSheet.FINAL_COLS
        }
        
        for period, columns in period_columns.items():
            assessment_no = 0
            for col in columns:
                max_score = self._as_optional_float(sheet[f"{col}{etl_const.OutputSheet.MAX_SCORE_ROW}"].value)
                if max_score is None: continue
                
                clo_code = self._as_optional_string(sheet[f"{col}{etl_const.OutputSheet.CLO_CODE_ROW}"].value)
                if clo_code is None: continue
                
                assessment_no += 1
                activity_name = self._as_optional_string(sheet[f"{col}{etl_const.OutputSheet.ACTIVITY_NAME_ROW}"].value)
                col_idx = column_index_from_string(col)
                
                for student in students:
                    raw_score = self._as_optional_float(sheet.cell(row=student["row"], column=col_idx).value)
                    records.append(RawScoreRecord(student_id=student["student_id"], student_name=student["student_name"] or "", grading_period=period, assessment_category="OUTPUT", assessment_no=assessment_no, clo_code=clo_code, activity_name=activity_name, max_score=max_score, raw_score=raw_score))
        return records

    def _merge_roster_by_name(self, sheet_students: list[dict[str, str | int | None]], db_students_by_name: dict[str, dict[str, str | int | None]]) -> list[dict[str, str | int | None]]:
        merged: list[dict[str, str | int | None]] = []
        for student in sheet_students:
            student_name = student.get("student_name")
            if not student_name: continue
            
            key = self._normalize_name(student_name)
            db_student = db_students_by_name.get(key)
            student_id = student.get("student_id") or (db_student.get("student_id") if db_student else None)
            
            merged.append({"student_id": student_id, "student_name": student_name, "row": student.get("row")})
        return merged

    def _find_cover_value(self, sheet: Worksheet, label: str) -> str | None:
        needle = self._normalize_label(label)
        for row in range(1, sheet.max_row + 1):
            for col in range(1, sheet.max_column + 1):
                value = sheet.cell(row=row, column=col).value
                if self._normalize_label(value) != needle: continue
                
                next_value = self._as_optional_string(sheet.cell(row=row, column=col + 1).value)
                return next_value
        return None

    @staticmethod
    def _normalize_label(value: Any) -> str:
        if value is None: return ""
        normalized = str(value).strip().upper()
        if normalized.endswith(":"): normalized = normalized[:-1]
        return " ".join(normalized.split())

    @staticmethod
    def _normalize_name(name: str) -> str: return " ".join(name.strip().lower().split())

    @staticmethod
    def _normalize_text(value: Any) -> str:
        if value is None: return ""
        return " ".join(str(value).strip().upper().split())

    @staticmethod
    def _as_optional_string(value: Any) -> str | None:
        if value is None: return None
        text = str(value).strip()
        return text if text else None

    def _as_string(self, value: Any) -> str:
        text = self._as_optional_string(value)
        return text or ""

    @staticmethod
    def _as_optional_float(value: Any) -> float | None:
        if value is None: return None
        if isinstance(value, str):
            stripped = value.strip()
            if not stripped: return None
            value = stripped
        try: return float(value)
        except (TypeError, ValueError): return None

    def _as_float(self, value: Any) -> float:
        parsed = self._as_optional_float(value)
        return parsed if parsed is not None else 0.0

    @staticmethod
    def _as_int(value: Any) -> int:
        if value is None: return 0
        if isinstance(value, str):
            stripped = value.strip()
            if not stripped: return 0
            value = stripped
        try: return int(float(value))
        except (TypeError, ValueError): return 0

    def _coalesce_optional(self, primary: Any, fallback: Any) -> str | None:
        return self._as_optional_string(primary) or self._as_optional_string(fallback)
