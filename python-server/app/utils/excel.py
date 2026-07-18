"""Generic Excel workbook helpers used by the ETL pipeline.

The helpers intentionally stay agnostic of any class-record structure.
`read_rows()` returns the raw worksheet grid as ``list[list[Any]]`` with the
first row representing the sheet header row. Later extraction steps can map
that grid into domain-specific records.
"""

from __future__ import annotations

import asyncio
from pathlib import Path
from typing import Any, Iterable
from zipfile import BadZipFile

from openpyxl import load_workbook
from openpyxl.utils.exceptions import InvalidFileException

from app.core.exceptions import InvalidWorkbook, MissingWorksheet


def _open_workbook(path: str | Path, *, data_only: bool = True):
    try:
        return load_workbook(filename=str(path), read_only=True, data_only=data_only)
    except (BadZipFile, FileNotFoundError, InvalidFileException, OSError) as exc:
        raise InvalidWorkbook(f"Unable to open workbook: {path}") from exc


def _read_sheet_rows_sync(
    path: str | Path,
    sheet_name: str,
    *,
    data_only: bool = True,
) -> list[list[Any]]:
    workbook = _open_workbook(path, data_only=data_only)
    try:
        if sheet_name not in workbook.sheetnames:
            raise MissingWorksheet(sheet_name)

        worksheet = workbook[sheet_name]
        return [list(row) for row in worksheet.iter_rows(values_only=True)]
    finally:
        workbook.close()


def _normalize_headers(headers: Iterable[Any]) -> list[str]:
    normalized = ["" if header is None else str(header).strip() for header in headers]
    while normalized and normalized[-1] == "":
        normalized.pop()
    return normalized


async def list_sheets(path: str | Path) -> list[str]:
    """Return the worksheet names contained in an Excel workbook."""

    workbook = await asyncio.to_thread(_open_workbook, path)
    try:
        return list(workbook.sheetnames)
    finally:
        workbook.close()


async def read_rows(path: str | Path, sheet_name: str) -> list[list[Any]]:
    """Return the raw worksheet grid for `sheet_name`.

    The return value is a list of rows, where each row is a list of cell values
    from `openpyxl`. The first row is the header row. Empty cells are preserved
    as `None` values so downstream parsing can distinguish missing data from
    blank strings.
    """

    return await asyncio.to_thread(_read_sheet_rows_sync, path, sheet_name)


async def validate_headers(
    path: str | Path,
    sheet_name: str,
    expected_headers: Iterable[Any],
) -> bool:
    """Ensure the first worksheet row matches the expected header structure.

    Comparison is strict after normalizing cell values to trimmed strings and
    removing trailing blank header cells.
    """

    rows = await read_rows(path, sheet_name)
    actual_headers = _normalize_headers(rows[0]) if rows else []
    expected = _normalize_headers(expected_headers)

    if actual_headers != expected:
        raise InvalidWorkbook(
            f"Header mismatch in '{sheet_name}': expected {expected}, got {actual_headers}"
        )

    return True

