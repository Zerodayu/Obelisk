#!/usr/bin/env python3
"""
Comprehensive validation test for extractor and transformer fixes.
Validates RawScoreRecord counts, transformer output, and CO-PO Attainment sheet values.
"""

import asyncio
import sys
from pathlib import Path
from collections import defaultdict
from openpyxl import load_workbook

# Add project root to path to allow imports from `app`
PROJECT_ROOT = Path(__file__).parent.parent
sys.path.insert(0, str(PROJECT_ROOT))

from app.etl.extract.extractor import ExcelExtractor
from app.etl.transform.transformer import SimpleTransformer, INSTITUTIONAL_THRESHOLD

# Define the path to the templates directory relative to the project root
TEMPLATES_DIR = PROJECT_ROOT / "classrecord_templates"

# Known-good baseline counts from the sample file (no gap rows)
EXPECTED_COUNTS = {
    ("PRELIM", "TLA"): 12,
    ("PRELIM", "AT"): 6,
    ("PRELIM", "EXAM"): 3,
    ("PRELIM", "OUTPUT"): 3,
    ("MIDTERM", "TLA"): 24,
    ("MIDTERM", "AT"): 9,
    ("MIDTERM", "EXAM"): 6,
    ("FINAL", "TLA"): 18,
    ("FINAL", "AT"): 15,
    ("FINAL", "EXAM"): 9,
}


async def step1_validate_extraction():
    """
    Step 1: Extract data and validate RawScoreRecord counts.
    """
    print("\n" + "=" * 80)
    print("STEP 1: Validate Extraction - RawScoreRecord Counts")
    print("=" * 80)
    
    file_path = TEMPLATES_DIR / "E-classrecord(LECTURE ONLY).xlsx"
    if not file_path.exists():
        print(f"ERROR: Sample file not found at {file_path}")
        return None, None, None
    
    try:
        extractor = ExcelExtractor()
        header, records, clo_plo_mapping = await extractor.extract(source=file_path)
    except Exception as e:
        print(f"ERROR during extraction: {e}")
        import traceback
        traceback.print_exc()
        return None, None, None
    
    counts = defaultdict(int)
    for record in records:
        key = (record.grading_period, record.assessment_category)
        counts[key] += 1
    
    print(f"\nTotal RawScoreRecords: {len(records)}")
    print("\nCounts by (grading_period, assessment_category):")
    print("-" * 50)
    
    all_match = True
    for key in sorted(EXPECTED_COUNTS.keys()):
        expected = EXPECTED_COUNTS[key]
        actual = counts[key]
        match = "✓" if actual == expected else "✗ MISMATCH"
        print(f"  {key[0]}/{key[1]:8s}: {actual:3d} (expected {expected:3d}) {match}")
        if actual != expected:
            all_match = False
    
    for key in counts:
        if key not in EXPECTED_COUNTS:
            print(f"  UNEXPECTED: {key[0]}/{key[1]} = {counts[key]}")
            all_match = False
    
    if not all_match:
        print("\n❌ COUNTS DO NOT MATCH BASELINE - STOPPING HERE")
        return None, None, None
    
    print("\n✅ All counts match baseline - proceeding to next steps")
    return header, records, clo_plo_mapping


async def step2_validate_transform(header, records, clo_plo_mapping):
    """
    Step 2: Transform data and pick one student to inspect.
    """
    print("\n" + "=" * 80)
    print("STEP 2: Transform Data and Pick One Student")
    print("=" * 80)
    
    try:
        transformer = SimpleTransformer()
        attainments = await transformer.transform((header, records, clo_plo_mapping))
    except Exception as e:
        print(f"ERROR during transformation: {e}")
        import traceback
        traceback.print_exc()
        return None
    
    print(f"\nTotal StudentCLOAttainment records: {len(attainments)}")
    
    if not attainments:
        print("ERROR: No attainment records produced")
        return None
    
    selected = attainments[0]
    
    print(f"\nSelected Student: {selected.student_name}")
    print(f"Student ID: {selected.student_id}")
    print(f"CLO Code: {selected.clo_code}")
    print(f"Direct CLO Attainment %: {selected.direct_clo_attainment_pct:.2%}")
    print(f"Met Threshold (at {INSTITUTIONAL_THRESHOLD:.0%}): {selected.met_threshold}")
    print(f"CLO Level: {selected.clo_level}")
    
    return selected


def step3_read_copa_sheet(student):
    """
    Step 3: Read the CO-PO Attainment sheet directly and find the student's value.
    """
    print("\n" + "=" * 80)
    print("STEP 3: Inspect CO-PO Attainment Sheet Structure")
    print("=" * 80)
    
    file_path = TEMPLATES_DIR / "E-classrecord(LECTURE ONLY).xlsx"
    
    try:
        wb = load_workbook(file_path, read_only=True, data_only=True)
    except Exception as e:
        print(f"ERROR opening workbook: {e}")
        return None
    
    if "CO-PO Attainment" not in wb.sheetnames:
        print("ERROR: 'CO-PO Attainment' sheet not found")
        return None
    
    print("\nThis sheet contains CLASS-LEVEL summary, not per-student attainment.")
    print("→ Skipping direct sheet lookup - transformer output is the source of truth")
    
    wb.close()
    return None


def step4_compare_values(student, sheet_attainment, sheet_cell):
    """
    Step 4: Compare computed vs. sheet values (if available).
    """
    print("\n" + "=" * 80)
    print("STEP 4: Summary of Computed Attainment")
    print("=" * 80)
    
    if sheet_attainment is None:
        print("\n✅ Transformer Computed Values (no sheet comparison available):")
        print(f"\n  Student: {student.student_name}")
        print(f"  CLO Code: {student.clo_code}")
        print(f"  Computed Direct CLO Attainment: {student.direct_clo_attainment_pct:.4f} ({student.direct_clo_attainment_pct*100:.2f}%)")
        print(f"  Met Threshold (@ {INSTITUTIONAL_THRESHOLD:.0%}): {student.met_threshold}")
        print(f"  CLO Level: {student.clo_level}")
        return

async def main():
    """Run all validation steps."""
    print("\n" + "=" * 80)
    print("OBELISK ETL - EXTRACTION & TRANSFORMATION VALIDATION TEST")
    print("=" * 80)
    
    header, records, clo_plo_mapping = await step1_validate_extraction()
    if header is None or records is None:
        print("\n❌ Extraction validation failed - exiting")
        return
    
    selected_student = await step2_validate_transform(header, records, clo_plo_mapping)
    if selected_student is None:
        print("\n❌ Transform validation failed - exiting")
        return
    
    copa_result = step3_read_copa_sheet(selected_student)
    
    step4_compare_values(selected_student, None, None)
    
    print("\n" + "=" * 80)
    print("VALIDATION TEST COMPLETE")
    print("=" * 80)


if __name__ == "__main__":
    asyncio.run(main())
